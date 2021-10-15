import csv
from openpyxl import Workbook
import sys
from datetime import datetime
import os
from collections import OrderedDict

def die(message):
    print(message)
    sys.exit(1)

# TODO:
# Flag if over 8 hours on a job
class TimeSheet(object):

    # High level dictionary to track employee work
    timeCards = OrderedDict()

    # TODO: Use schema to validate dictionary

    # employeeName -> dict_of_workdays -> jobs

    # Constructor
    def __init__(self):
        pass

    def getEmployeeTimeCard(self, name):
        return self.timeCards[name]

    def getTotalWorkedHours(self, name):
        totalMins = 0
        tc = self.getEmployeeTimeCard(name)
        for workDay, jobList in tc.items():
            for jobs in jobList:
                for minsWorked in jobs:
                    totalMins += (jobs[minsWorked])
        return totalMins

# Function to pretty print dictionary
def pretty(d, indent=0):
   for key, value in d.items():
      print('\t' * indent + str(key))
      if isinstance(value, dict):
         pretty(value, indent+1)
      else:
         print('\t' * (indent+1) + str(value))


# Parsing function to enable sorting;
# if clock_in data is an empty string, treat it as 1/1/2020
def getDate(x):
        if x == '':
            return '1/1/2020'
        else:
            return x

def main():

    # Create instance of TimeSheet class to track per employee TimeCards
    timeSheet = TimeSheet();

    uploadedFilename = sys.argv[1]

    # Build path to uploaded file. Node passes name of uploaded file as first argument
    fileToOpen = 'routes/uploads/' + uploadedFilename

    if not os.path.isfile(fileToOpen):
        # Check for file in expected location. i.e. routes/uploads/<filename>
        die('Upload failed {}' .format(uploadedFilename))
    else:
        # Found file at expect path
        # basic 'check' for csv
        if not uploadedFilename.split('.')[-1] == 'csv':
            die('Expected csv file but got {} file' .format(uploadedFilename.split('.')[-1]))

    # list of keys python expects in csv
    csvKeys = ['tech', 'clock_in', 'job_number', 'total_time_minutes', 'company']

    with open(fileToOpen, encoding="utf8", mode='r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        # Sort CSV based on date
        #sorted_csv = sorted(csv_reader, key=lambda row: (datetime.strptime(row['clock_in'].split(' ')[0], "%m/%d/%Y")))
        sorted_csv = sorted(csv_reader, key=lambda row: getDate(row['clock_in']))
        line_count = 0
        for row in sorted_csv:

            # Check key's if they exists
            # Assume if keys are in row 0, they are in all rows
            if (line_count == 0):
                missingKeys = ''
                for searchKey in csvKeys:
                    if not searchKey in row.keys():
                        missingKeys += searchKey + ' '

                if missingKeys:
                    die('Uploaded CSV missing the following columns: {}' .format(missingKeys))

            employeeName = row['tech']
            # check if blank, assign default name
            if not employeeName:
                employeeName = "missing_name"

            day = row['clock_in']
            if not day:
                day = "1/1/2020"
            day = day.split(' ')
            day = day[0]


            companyId = row['company']
            if not companyId:
                companyId = "missing_company_id"

            jobId = row['job_number']
            if not jobId:
                jobId = "missing_jobid"

            jobId = companyId + ' ' + jobId        

            minsWorked = row['total_time_minutes']
            if not minsWorked:
                minsWorked = 0
            else:
                minsWorked = int(minsWorked)/60

            if timeSheet.timeCards.get(employeeName) == None:
                job = {jobId: minsWorked}
                jobList = [job]
                workDate = {day: jobList}
                timeSheet.timeCards.update({employeeName: workDate})
            else:
                # Timecard exists, check if there is worked logged for this date
                if timeSheet.timeCards[employeeName].get(day) == None:
                    job = {jobId: minsWorked}
                    jobList = [job]
                    timeSheet.timeCards[employeeName].update({day: jobList})
                else:
                    jobList = timeSheet.timeCards[employeeName][day]
                    # Check for matching jobIds to combined into a single entry
                    list_index = 0
                    for jobs in jobList:
                        job_id_found = False
                        for job_id, mins_worked in jobs.items():
                            if (jobId == job_id):
                                job_id_found = True
                                # update existing entry
                                job = {jobId : (mins_worked + minsWorked)}
                                jobList[list_index] = job
                                timeSheet.timeCards[employeeName].update({day: jobList})
                                break
                        list_index += 1
                    if (job_id_found == False):
                        job = {jobId: minsWorked}
                        jobList.append(job)
                        # Overwrite existing date with new job info
                        timeSheet.timeCards[employeeName].update({day: jobList})

            line_count += 1

    #pretty(timeSheet.timeCards)

    # Create xlsx file with a sheet for each employee
    workbook = Workbook()

    # Delete default sheet
    defaultSheet=workbook['Sheet']
    workbook.remove(defaultSheet)

    employee_jobid_to_row_lut = {}

    for employee in timeSheet.timeCards:
        activeSheet = workbook.create_sheet(employee)
        activeSheet.cell(column=1, row=1, value="Employee Name:")
        # Get Column String from column int
        #activeSheet.column_dimensions[str(chr(64 + 1))].width = 14
        activeSheet.cell(column=2, row=1, value=employee)
        activeSheet.cell(column=1, row=2, value="Job ID:")
        col = 2
        row = 4
        for day in timeSheet.timeCards[employee]:
            # Write Dates across columns
            activeSheet.cell(column=col, row=2, value=day)
            # Get Column String from column int
            #activeSheet.column_dimensions[str(chr(64 + col))].width = 10
            for jobDict in timeSheet.timeCards[employee][day]:
                for jobId, minsWorked in jobDict.items():
                    # Check if employee has worked at this jobid preivously,
                    # if so, write minsWorked on same row
                    temp = row;
                    inc = 1

                    # Check if this employee has an entry in our lookup table
                    if employee_jobid_to_row_lut.get(employee) == None:
                        jobid_to_row = {jobId: row}
                        employee_jobid_to_row = {employee: jobid_to_row}
                        employee_jobid_to_row_lut.update(employee_jobid_to_row)

                    # Check if employee already has worked on this job    
                    elif employee_jobid_to_row_lut[employee].get(jobId) == None:
                        employee_jobid_to_row_lut[employee].update({jobId: row})

                    # Employee worked at this job before. Lookup row number
                    else:
                        row = employee_jobid_to_row_lut[employee][jobId];
                        inc = 0

                    # Write Job ID
                    activeSheet.cell(column=1, row=row, value=jobId)
                    activeSheet.cell(column=col, row=row, value=round(minsWorked, 2))
                    row = temp
                    row += inc
            col += 1
        col -= 2 
        row += 5
        activeSheet.cell(column=col, row=row, value="Total Hours:")
        col += 1
        activeSheet.cell(column=col, row=row, value=(timeSheet.getTotalWorkedHours(employee)))

    # datetime object containing current date and time
    now = datetime.now()

    # d_m_y_h_m_s
    dt_string = now.strftime("%d_%m_%Y_%H_%M_%S")

    # Build filename/path
    filePath = 'routes/downloads/'
    fileName ='timeCards_' + dt_string + '.xlsx'
    fileDest = filePath+fileName

    # pipe filename back to node js via standard output
    print(fileName)

    # Create directory if it doesn't exist
    if not os.path.exists(os.path.join(os.getcwd(), filePath)) :
        os.mkdir(os.path.join(os.getcwd(), filePath))
        
    # Save file
    workbook.save(filename=fileDest)

    # return 'success'
    sys.exit(0)

if __name__ == "__main__":
    main()